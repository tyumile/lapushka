#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
make_template_and_expected.py

Назначение:
- Берёт "заполненный" Excel ИД (как эталон).
- Делает из него ШАБЛОН:
  * скалярные значения, которые должны подставляться -> заменяет на {{field_key}}
  * табличные части (реестры/ведомости с заголовком "№ п/п") -> выносит строки в JSON.tables,
    добавляет якорь {{TABLE:table_key}} и очищает значения в строках таблицы (НЕ ломая стиль).
- Генерирует эталонный JSON ответа агента (expected): values + tables + tables_contract + occurrences.
- Генерирует промпт для агента-проверяльщика.

ВАЖНО:
- Скрипт бережно относится к стилям: НЕ трогает границы/шрифты/мерджи/форматы.
- При очистке табличных строк НЕ пишет в MergedCell (иначе openpyxl падает).
"""

import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell.cell import MergedCell


# ---------------------------
# Общие утилиты
# ---------------------------

RU_MAP = {
    'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e', 'ё': 'e', 'ж': 'zh', 'з': 'z', 'и': 'i', 'й': 'y',
    'к': 'k', 'л': 'l', 'м': 'm', 'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r', 'с': 's', 'т': 't', 'у': 'u', 'ф': 'f',
    'х': 'h', 'ц': 'ts', 'ч': 'ch', 'ш': 'sh', 'щ': 'sch', 'ъ': '', 'ы': 'y', 'ь': '', 'э': 'e', 'ю': 'yu', 'я': 'ya'
}


def translit(s: str) -> str:
    out = []
    for ch in (s or "").lower():
        if ch in RU_MAP:
            out.append(RU_MAP[ch])
        elif ch.isalnum():
            out.append(ch)
        else:
            out.append('_')
    return re.sub(r"_+", "_", "".join(out)).strip("_")


def norm(v):
    if v is None:
        return None
    if isinstance(v, str):
        return v.strip()
    return v


def is_formula_value(v) -> bool:
    return isinstance(v, str) and v.startswith("=")


# ---------------------------
# Таблицы (реестры)
# ---------------------------

PP_RE = re.compile(r"№\s*п/п", re.IGNORECASE)


def find_table_header(ws, max_rows=300, max_cols=60):
    """
    Ищет в листе заголовок таблицы по маркеру "№ п/п".
    Возвращает (header_row, header_col) или None.
    """
    for r in range(1, min(ws.max_row, max_rows) + 1):
        for c in range(1, min(ws.max_column, max_cols) + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and PP_RE.search(v.strip()):
                return (r, c)
    return None


def detect_columns(ws, header_row, start_col=1, end_col=60):
    """
    Пытается сопоставить заголовки колонок стандартным ключам.
    Возвращает dict: column_key -> column_letter
    """
    mapping = {}
    for c in range(start_col, min(end_col, ws.max_column) + 1):
        v = ws.cell(header_row, c).value
        if not isinstance(v, str):
            continue
        t = re.sub(r"\s+", " ", v.strip().lower())

        if PP_RE.search(t):
            mapping["pp"] = get_column_letter(c)
        elif "наимен" in t and "документ" in t:
            mapping["doc_name"] = get_column_letter(c)
        elif "№" in t and "дата" in t:
            mapping["doc_no_date"] = get_column_letter(c)
        elif "организац" in t and ("состав" in t or "выдав" in t or "составив" in t):
            mapping["issuer"] = get_column_letter(c)
        elif ("кол-во" in t or "количество" in t) and "лист" in t:
            mapping["sheets_count"] = get_column_letter(c)
        elif "лист" in t and "№" in t:
            mapping["sheet_no"] = get_column_letter(c)
        elif "примеч" in t:
            mapping["note"] = get_column_letter(c)

    return mapping


def find_last_table_row(ws, header_row, col_letters, blank_limit=5):
    """
    Определяет последнюю строку таблицы по "полосе пустых строк" в ключевых колонках.
    """
    last = header_row
    blank = 0
    for r in range(header_row + 1, ws.max_row + 1):
        vals = [norm(ws[f"{cl}{r}"].value) for cl in col_letters]
        if all(v in (None, "") for v in vals):
            blank += 1
            if blank >= blank_limit:
                break
        else:
            blank = 0
            last = r
    return last


def place_table_anchor_far_right(ws, table_key: str):
    """
    Ставит якорь {{TABLE:<table_key>}} в строке 1 в первой пустой ячейке справа (чтобы не ломать верстку).
    """
    r = 1
    c = ws.max_column + 1
    while c < ws.max_column + 150 and ws.cell(r, c).value not in (None, ""):
        c += 1
    anchor_cell = f"{get_column_letter(c)}{r}"
    anchor_token = f"{{{{TABLE:{table_key}}}}}"
    ws[anchor_cell].value = anchor_token
    return anchor_cell, anchor_token


def safe_clear_cell(ws, row_idx: int, col_letter: str):
    """
    Очищает значение в ячейке таблицы, НЕ трогая MergedCell (иначе openpyxl упадёт).
    """
    col_idx = column_index_from_string(col_letter)
    cell = ws.cell(row_idx, col_idx)
    if isinstance(cell, MergedCell):
        # Внутри merge-диапазона value readonly. Не трогаем.
        return
    cell.value = ""


# ---------------------------
# Скалярные значения (плейсхолдеры)
# ---------------------------

DATE_RE = re.compile(r"\b\d{2}\.\d{2}\.\d{4}\b")
PROJ_CODE_RE = re.compile(r"\b\d{2,4}-\d{1,2}-\d{1,2}-[A-Za-zА-Яа-я0-9.]+(?:\.\d+)?\b")
ACT_NO_RE = re.compile(r"№\s*\d+")
INLINE_LABEL_VALUE_RE = re.compile(r".+:\s*\S+")
LABEL_COLON_RE = re.compile(r".+:\s*$")

STATIC_RE = re.compile(
    r"(№\s*п/п|№\s*и\s*дата|Организация, составившая|Кол-во\s*лист|Лист\s*№|"
    r"ИСПОЛНИТЕЛЬНАЯ\s+ДОКУМЕНТАЦИЯ|РАБОЧАЯ\s+ДОКУМЕНТАЦИЯ|"
    r"Тюмень,\s*20\d{2}\s*г|"
    r"обществ(о|а) с ограниченной ответственностью|ИНН|КПП|БИК|тел|mail)"
    , re.IGNORECASE
)


def infer_label(ws, r, c):
    # слева
    for dc in range(1, 10):
        cc = c - dc
        if cc <= 0:
            break
        v = ws.cell(r, cc).value
        if isinstance(v, str):
            t = v.strip()
            if not t:
                continue
            if LABEL_COLON_RE.fullmatch(t):
                return t.rstrip(":").strip()
            if len(t) <= 60 and any(k in t.lower() for k in ["объект", "адрес", "шифр", "обознач", "стад", "корпус", "заказчик", "проектиров", "дата", "№"]):
                return t.rstrip(":").strip()

    # сверху
    for dr in range(1, 8):
        rr = r - dr
        if rr <= 0:
            break
        v = ws.cell(rr, c).value
        if isinstance(v, str):
            t = v.strip()
            if not t:
                continue
            if len(t) <= 60 and any(k in t.lower() for k in ["объект", "адрес", "шифр", "обознач", "стад", "корпус", "заказчик", "проектиров", "дата", "№"]):
                return t.rstrip(":").strip()

    return None


def canonical_key(label: str, sheet_title: str):
    if not label:
        return None
    l = label.lower()
    if "наимен" in l and "объект" in l:
        return "object_name"
    if "адрес" in l:
        return "object_address"
    if "шифр" in l or "обознач" in l:
        return "project_code"
    if "стад" in l:
        return "stage"
    if "корпус" in l or "секци" in l:
        return "building_part"
    if "заказчик" in l:
        return "customer_org"
    if "проектиров" in l:
        return "designer_org"
    if "дата" in l and ("аоср" in sheet_title.lower() or "аоок" in sheet_title.lower() or "акт" in sheet_title.lower()):
        return "act_date"
    if ("№" in label or "номер" in l) and ("аоср" in sheet_title.lower() or "аоок" in sheet_title.lower() or "акт" in sheet_title.lower()):
        return "act_no"
    return None


def is_variable_value(v):
    if v is None:
        return False

    if isinstance(v, str):
        s = v.strip()
        if not s:
            return False
        if STATIC_RE.search(s):
            return False
        if INLINE_LABEL_VALUE_RE.fullmatch(s):
            # "Дата: 14.10.2025" оставляем как статичный текст шаблона
            return False
        if s.upper() == s and len(s) > 12 and any(ch.isalpha() for ch in s):
            return False

        # вероятно переменные значения
        if DATE_RE.search(s):
            return True
        if ACT_NO_RE.search(s):
            return True
        if PROJ_CODE_RE.search(s):
            return True
        if "ул." in s.lower() and any(ch.isdigit() for ch in s):
            return True
        if "этап" in s.lower() and any(ch.isdigit() for ch in s):
            return True
        if "корпус" in s.lower() and any(ch.isdigit() for ch in s):
            return True
        if any(k in s for k in ["ООО", "АО", "ИП", "ПАО", "ЗАО"]):
            return True
        if len(s) > 40 and ("комплекс" in s.lower() or "жил" in s.lower()):
            return True

        return False

    if isinstance(v, (int, float)):
        return True

    return False


# ---------------------------
# Main
# ---------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--filled", required=True, help="Заполненный Excel ИД (эталон)")
    ap.add_argument("--project", default="", help="Проект PDF (опционально)")
    ap.add_argument("--out_prefix", required=True, help="Префикс выходных файлов, например EX1_KJ_Foundation")
    args = ap.parse_args()

    filled = Path(args.filled)
    project = Path(args.project) if args.project else None
    out_prefix = args.out_prefix

    wb = load_workbook(filled, data_only=False)

    # Подсветка (не обязательно; при желании можно убрать)
    fill_scalar = PatternFill("solid", fgColor="DDEBF7")  # light blue
    fill_anchor = PatternFill("solid", fgColor="FFF2CC")  # light yellow

    tables_contract = {}
    tables_rows = {}

    scalar_values = {}
    occurrences = {}

    used_keys = set()

    def make_key(label, sheet, coord):
        base = canonical_key(label, sheet) or translit(label) or translit(f"{sheet}_{coord}")
        key = base or translit(f"{sheet}_{coord}")
        i = 2
        while key in used_keys:
            key = f"{base}_{i}"
            i += 1
        used_keys.add(key)
        return key

    # 1) Find table regions
    table_regions = {}
    for ws in wb.worksheets:
        hdr = find_table_header(ws)
        if not hdr:
            continue

        header_row, header_col = hdr
        colmap = detect_columns(ws, header_row, 1, 60)
        if not colmap:
            continue

        last_row = find_last_table_row(ws, header_row, list(colmap.values()), blank_limit=5)
        table_regions[ws.title] = {"header_row": header_row, "last_row": last_row, "colmap": colmap}

    # 2) Process tables: anchor + extract rows + clear values
    for sheet_name, reg in table_regions.items():
        ws = wb[sheet_name]
        table_key = translit(sheet_name) or sheet_name

        anchor_cell, anchor_token = place_table_anchor_far_right(ws, table_key)
        ws[anchor_cell].fill = fill_anchor

        header_row = reg["header_row"]
        colmap = reg["colmap"]

        rows = []
        for rr in range(header_row + 1, reg["last_row"] + 1):
            vals = {k: norm(ws[f"{letter}{rr}"].value) for k, letter in colmap.items()}
            if all(v in (None, "") for v in vals.values()):
                continue

            rows.append(vals)

            # очищаем значения в строке таблицы, не трогая merges
            for letter in colmap.values():
                safe_clear_cell(ws, rr, letter)

        tables_contract[table_key] = {
            "sheet_name": sheet_name,
            "anchor_token": anchor_token,
            "anchor_cell": anchor_cell,
            "header_row": header_row,
            "template_row_index": header_row + 1,
            "columns": colmap
        }
        tables_rows[table_key] = rows

    # 3) Process scalar placeholders (skip table header+body in mapped columns)
    for ws in wb.worksheets:
        reg = table_regions.get(ws.title)
        header_row = reg["header_row"] if reg else None
        last_row = reg["last_row"] if reg else None
        mapped_cols = set(reg["colmap"].values()) if reg else set()

        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if v is None or is_formula_value(v):
                    continue

                # skip table header row
                if header_row and cell.row == header_row:
                    continue
                # skip table body cells in mapped columns
                if header_row and last_row and header_row < cell.row <= last_row:
                    if get_column_letter(cell.column) in mapped_cols:
                        continue

                if not is_variable_value(v):
                    continue

                label = infer_label(ws, cell.row, cell.column)
                key = canonical_key(label, ws.title) or make_key(label, ws.title, cell.coordinate)
                placeholder = "{{" + key + "}}"

                # если ключ уже есть, но значение другое — делаем новый ключ
                if key in scalar_values and scalar_values[key] != norm(v):
                    key = make_key(label or key, ws.title, cell.coordinate)
                    placeholder = "{{" + key + "}}"

                scalar_values[key] = norm(v)
                occurrences.setdefault(key, []).append({"sheet": ws.title, "cell": cell.coordinate})

                cell.value = placeholder
                cell.fill = fill_scalar

    # 4) Save outputs
    out_template = Path(f"{out_prefix}_template.xlsx")
    wb.save(out_template)

    expected = {
        "meta": {
            "example_id": out_prefix,
            "project_pdf": project.name if project else "",
            "filled_excel": filled.name,
            "template_excel": out_template.name
        },
        "values": scalar_values,
        "tables": tables_rows,
        "tables_contract": tables_contract,
        "occurrences": occurrences
    }

    out_expected = Path(f"{out_prefix}_expected.json")
    out_expected.write_text(json.dumps(expected, ensure_ascii=False, indent=2), encoding="utf-8")

    reviewer_prompt = f"""ПРОВЕРКА ШАБЛОНА И ЭТАЛОНА

Вход:
- проект: {project.name if project else "(не указан)"}
- заполненный ИД: {filled.name}
- шаблон: {out_template.name}
- эталон JSON: {out_expected.name}

Проверь:
1) Шаблон не меняет стиль/верстку заполненного ИД (границы, шрифты, объединения, форматы).
2) Плейсхолдеры {{...}} стоят только там, где должны подставляться проектные/экземплярные значения.
   Статичные заголовки/подписи/шапки не должны становиться плейсхолдерами.
3) Таблицы:
   - якорь {{TABLE:<table_key>}} есть (см. tables_contract)
   - строки из expected.json tables соответствуют заполненному ИД
   - columns-map корректно отражает колонки таблицы
4) Если применить expected.json к шаблону — должен получиться заполненный ИД по содержимому.

Ответ:
- Ошибки (sheet/cell или table_key) + как исправить
- Пропуски (что нужно было заменить, но не заменено)
- Лишнее (что заменено, но должно быть статичным)
"""
    out_review = Path(f"{out_prefix}_reviewer_prompt.txt")
    out_review.write_text(reviewer_prompt, encoding="utf-8")

    print("OK")
    print("template:", out_template)
    print("expected json:", out_expected)
    print("reviewer prompt:", out_review)


if __name__ == "__main__":
    main()
