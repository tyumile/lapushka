import json
import os
import secrets
import shutil
from datetime import datetime
from pathlib import Path

from ..llm.prompt_loader_v02 import load_prompt
from ..llm.responses_client_v02 import ResponsesClientV02
from .dictionary_service import get_razdel, load_dictionary
from .llm_runtime import SUPPORTED_CONTEXT_EXTS
from .project_storage import persist_run_artifact, persist_run_json, project_root, save_processing_json

EXCLUDED_REL = "01_input/99_excluded"
BINARY_UPLOAD_DENY_EXTS = {".xls", ".xlsx", ".xlsm"}
FILE_UPLOAD_ALLOWED_EXTS = {".pdf", ".txt", ".md", ".json", ".png", ".jpg", ".jpeg"}


def _move_to_excluded(root: Path, path: Path, reason: str) -> Path:
    try:
        rel = path.relative_to(root)
    except ValueError:
        rel = Path(path.name)
    rel_str = str(rel).replace("\\", "/")
    if rel_str.startswith(EXCLUDED_REL):
        return path
    excluded_dir = root / EXCLUDED_REL
    excluded_dir.mkdir(parents=True, exist_ok=True)
    base = rel.name
    dest = excluded_dir / base
    idx = 1
    while dest.exists() and dest != path:
        stem, suf = (dest.stem, dest.suffix) if dest.suffix else (dest.name, "")
        dest = excluded_dir / f"{stem}_{idx}{suf}"
        idx += 1
    if path != dest:
        shutil.move(str(path), str(dest))
    meta_path = dest.with_suffix(dest.suffix + ".excluded_reason.txt")
    meta_path.write_text(reason, encoding="utf-8")
    return dest


def _collect_input_files(project_id: str, razdel_code: str) -> list[Path]:
    root = project_root(project_id)
    excl_prefix = (root / EXCLUDED_REL).as_posix()

    def _skip_excluded(p: Path) -> bool:
        return not p.as_posix().startswith(excl_prefix)

    files: list[Path] = []
    files.extend([p for p in (root / "01_input" / "01_project").rglob("*") if p.is_file() and _skip_excluded(p)])
    files.extend([p for p in (root / "01_input" / "03_ojr").rglob("*") if p.is_file() and _skip_excluded(p)])

    qreg = root / "02_processing" / "p2_quality_registry_final.json"
    if qreg.exists():
        files.append(qreg)
    dict_path = Path(__file__).resolve().parent.parent / "data" / "id_dictionary_v02.json"
    if dict_path.exists():
        files.append(dict_path)

    regs_dir = root / "06_regs" / razdel_code
    for reg_name in ("rules.md", "rules.json"):
        p = regs_dir / reg_name
        if p.exists():
            files.append(p)

    sample_dirs = [
        root / "01_input" / "04_samples" / razdel_code / "projects",
        root / "01_input" / "04_samples" / razdel_code / "id",
        root / "01_input" / "04_samples" / razdel_code / "doc_types",
    ]
    for d in sample_dirs:
        files.extend([p for p in d.rglob("*") if p.is_file() and _skip_excluded(p)])
    return [p for p in files if p.suffix.lower() in SUPPORTED_CONTEXT_EXTS and _skip_excluded(p)]


def _read_project_comment(project_id: str) -> str:
    meta_path = project_root(project_id) / "05_project_meta.json"
    if not meta_path.exists():
        return ""
    try:
        payload = json.loads(meta_path.read_text(encoding="utf-8"))
        return (payload.get("comment") or "").strip()
    except json.JSONDecodeError:
        return ""


def _xlsx_to_json_text(path: Path) -> str:
    try:
        from openpyxl import load_workbook
    except Exception:
        return json.dumps({"file": path.name, "error": "openpyxl_missing"}, ensure_ascii=False)

    wb = load_workbook(path, data_only=True, read_only=True)
    sheets = []
    for ws in wb.worksheets:
        rows = []
        max_rows = 120
        max_cols = 40
        for ridx, row in enumerate(ws.iter_rows(max_row=max_rows, max_col=max_cols, values_only=True), start=1):
            row_vals = ["" if c is None else str(c) for c in row]
            if any(v.strip() for v in row_vals):
                rows.append({"row": ridx, "values": row_vals})
        sheets.append({"sheet": ws.title, "rows": rows})
    return json.dumps({"file": path.name, "sheets": sheets}, ensure_ascii=False)


def _build_request_inputs(project_id: str, razdel_code: str, input_files: list[Path]) -> tuple[list[Path], list[str], dict[str, str]]:
    root = project_root(project_id)
    upload_paths: list[Path] = []
    input_texts: list[str] = []
    skip_reasons: dict[str, str] = {}

    for p in input_files:
        rel = str(p.relative_to(root)).replace("\\", "/") if root in p.parents else p.name
        ext = p.suffix.lower()

        if ext in BINARY_UPLOAD_DENY_EXTS:
            if ext in {".xlsx", ".xlsm"}:
                input_texts.append(f"sample_xlsx_json::{rel}\n{_xlsx_to_json_text(p)}")
            else:
                input_texts.append(json.dumps({"file": rel, "warning": "xls_binary_not_supported"}, ensure_ascii=False))
            skip_reasons[rel] = "converted_to_input_text"
            continue

        # JSON registries/dictionary explicitly as input_text
        if rel.endswith("02_processing/p2_quality_registry_final.json"):
            input_texts.append(f"quality_registry_json::{rel}\n{p.read_text(encoding='utf-8', errors='ignore')}")
            skip_reasons[rel] = "sent_as_input_text"
            continue
        if rel.endswith("core_v02/data/id_dictionary_v02.json") or p.name == "id_dictionary_v02.json":
            input_texts.append(f"id_dictionary_json::{rel}\n{p.read_text(encoding='utf-8', errors='ignore')}")
            skip_reasons[rel] = "sent_as_input_text"
            continue

        # text journal/rules as input_text
        if ext in {".txt", ".md", ".json"}:
            input_texts.append(f"context_text::{rel}\n{p.read_text(encoding='utf-8', errors='ignore')}")
            skip_reasons[rel] = "sent_as_input_text"
            continue

        if ext in FILE_UPLOAD_ALLOWED_EXTS:
            upload_paths.append(p)
        else:
            skip_reasons[rel] = "unsupported_for_upload"

    # add explicit dictionary from package as input_text even if not in collected list
    dict_path = Path(__file__).resolve().parent.parent / "data" / "id_dictionary_v02.json"
    if dict_path.exists():
        input_texts.append(f"id_dictionary_json::core_v02/data/id_dictionary_v02.json\n{dict_path.read_text(encoding='utf-8', errors='ignore')}")

    # include selected regs as input_text
    regs_dir = root / "06_regs" / razdel_code
    for reg_name in ("rules.md", "rules.json"):
        p = regs_dir / reg_name
        if p.exists() and p.is_file():
            rel = str(p.relative_to(root)).replace("\\", "/")
            input_texts.append(f"regs::{rel}\n{p.read_text(encoding='utf-8', errors='ignore')}")

    return upload_paths, input_texts, skip_reasons


def _has_evidence_file(evidence) -> bool:
    if isinstance(evidence, dict):
        return bool((evidence.get("file") or "").strip())
    return False


def _validate_fields_object(fields: dict) -> bool:
    if not isinstance(fields, dict):
        return False
    for _, f in fields.items():
        if not isinstance(f, dict):
            return False
        if not all(k in f for k in ("value", "style_source", "evidence")):
            return False
    return True


def validate_p4_plan_payload(payload: dict) -> tuple[bool, str]:
    if not isinstance(payload, dict):
        return False, "payload must be object"
    for key in ("process", "razdel_code", "work_breakdown", "doc_instances", "questions"):
        if key not in payload:
            return False, f"missing key: {key}"
    if payload.get("process") != "p4_plan_docs":
        return False, "process must be p4_plan_docs"
    if not isinstance(payload.get("work_breakdown"), list):
        return False, "work_breakdown must be list"
    if not isinstance(payload.get("doc_instances"), list):
        return False, "doc_instances must be list"
    if not isinstance(payload.get("questions"), list):
        return False, "questions must be list"

    for i, w in enumerate(payload.get("work_breakdown") or []):
        if not isinstance(w, dict):
            return False, f"work_breakdown[{i}] must be object"
        if (w.get("status") == "ok") and not _has_evidence_file(w.get("evidence")):
            return False, f"work_breakdown[{i}].status=ok requires evidence.file"

    for i, d in enumerate(payload.get("doc_instances") or []):
        if not isinstance(d, dict):
            return False, f"doc_instances[{i}] must be object"
        if (d.get("status") == "ok") and not _has_evidence_file(d.get("evidence")):
            return False, f"doc_instances[{i}].status=ok requires evidence.file"
        if not _validate_fields_object(d.get("fields") if isinstance(d.get("fields"), dict) else {}):
            return False, f"doc_instances[{i}].fields invalid"
        for f_key, f_val in (d.get("fields") or {}).items():
            if (d.get("status") == "ok" or f_val.get("value") not in (None, "")) and not _has_evidence_file(f_val.get("evidence")):
                if d.get("status") == "ok":
                    return False, f"doc_instances[{i}].fields.{f_key} requires evidence.file"

    return True, "ok"


def _normalize_doc_instance(inst: dict, idx: int) -> dict:
    out = dict(inst) if isinstance(inst, dict) else {}
    out.setdefault("instance_id", out.get("instance_key") or f"i{idx}")
    out.setdefault("doc_id", "")
    out.setdefault("doc_variant_id", None)
    out.setdefault("doc_type_id", "")
    out.setdefault("doc_type_name", "")
    out.setdefault("doc_name", "")
    out.setdefault("doc_number", "")
    out.setdefault("multiplicity", "single")
    out.setdefault("basis", "from_work_requirement")
    out.setdefault("instance_key", out.get("instance_id") or f"i{idx}")
    out.setdefault("linked_work_type_ids", [])
    out.setdefault("scope", {})
    out.setdefault("fields", {})
    out.setdefault("status", out.get("overall_status") or "needs_extraction")
    out.setdefault("evidence", {"file": "", "page": None, "snippet": None})

    # backward-compatible aliases
    out["multi"] = out.get("multiplicity") == "multi"
    multiplier = {
        "axis": (out.get("scope") or {}).get("axes") or "",
        "label": (out.get("scope") or {}).get("section") or "",
        "confidence": 0,
    }
    out["multiplier"] = multiplier
    out["work_scope"] = out.get("work_scope") if isinstance(out.get("work_scope"), list) else []
    out["overall_status"] = out.get("status")
    out.setdefault("user_note", "")
    return out


def _normalize_payload(payload: dict, razdel_code: str, selected_doc_type_ids: list[str]) -> dict:
    out = payload if isinstance(payload, dict) else {}
    out.setdefault("process", "p4_plan_docs")
    out.setdefault("razdel_code", razdel_code)
    out.setdefault("detected_work_group_ids", [])
    out.setdefault("work_breakdown", [])
    out.setdefault("doc_instances", [])
    out.setdefault("registry_links", [])
    out.setdefault("questions", [])
    out.setdefault("warnings", [])
    out.setdefault("selected_doc_type_ids", selected_doc_type_ids)

    if not isinstance(out["doc_instances"], list):
        out["doc_instances"] = []
    out["doc_instances"] = [_normalize_doc_instance(inst, i) for i, inst in enumerate(out["doc_instances"])]

    # backward compatibility for old UI keys
    out["open_questions"] = [q.get("message") if isinstance(q, dict) else str(q) for q in (out.get("questions") or [])]
    out["issues"] = out.get("warnings") or []
    return out


def _mock_payload(razdel_code: str, selected_doc_type_ids: list[str]) -> dict:
    return {
        "process": "p4_plan_docs",
        "razdel_code": razdel_code,
        "detected_work_group_ids": ["kj_foundation"],
        "work_breakdown": [
            {
                "work_group_id": "kj_foundation",
                "work_type_id": "kj_rebar",
                "work_name": "Армирование фундаментной плиты",
                "status": "needs_extraction",
                "evidence": {"file": "", "page": None, "snippet": None},
            }
        ],
        "doc_instances": [
            {
                "doc_id": "AOSR",
                "doc_variant_id": "AOSR_REBAR",
                "doc_type_id": selected_doc_type_ids[0] if selected_doc_type_ids else None,
                "doc_type_name": "Акты скрытых работ",
                "doc_name": "Акт освидетельствования скрытых работ",
                "basis": "from_work_requirement",
                "instance_key": "AOSR_REBAR_1",
                "multiplicity": "multi",
                "linked_work_type_ids": ["kj_rebar"],
                "scope": {"object_part": None, "section": None, "floor": None, "axes": None, "capture": None, "stage": None, "lot": None, "notes": None},
                "fields": {
                    "field_presented_work": {
                        "value": "Армирование фундаментной плиты",
                        "style_source": "standard",
                        "evidence": {"file": "01_input/01_project/mock.pdf", "page": 1, "snippet": "армирование"},
                    }
                },
                "status": "needs_extraction",
                "evidence": {"file": "", "page": None, "snippet": None},
            }
        ],
        "registry_links": [],
        "questions": [{"id": "q1", "type": "missing_source", "message": "Уточните лист подтверждения работ", "related_doc_instance_key": None, "related_work_type_id": "kj_rebar"}],
        "warnings": [],
    }


def _load_p4_prompt() -> str:
    root = Path(__file__).resolve().parents[2]
    preferred = root / "core" / "llm" / "prompts" / "04_p4_plan_docs.txt"
    if preferred.exists():
        return preferred.read_text(encoding="utf-8")
    return load_prompt("04_p4_plan_docs")


def run_process_p4b_build_doc_plan(project_id: str, razdel_code: str, selected_doc_type_ids: list[str]) -> tuple[str, dict, list[dict]]:
    run_id = datetime.utcnow().strftime("%Y%m%d%H%M%S") + "-" + secrets.token_hex(3)
    model = "gpt-4.1-mini"
    root = project_root(project_id)
    input_files = _collect_input_files(project_id, razdel_code)
    razdel = get_razdel(razdel_code)
    user_comment = _read_project_comment(project_id)

    api_key = ""
    try:
        from django.conf import settings

        api_key = getattr(settings, "OPENAI_API_KEY", "") or ""
    except Exception:
        pass
    if not api_key:
        api_key = os.getenv("OPENAI_API_KEY", "") or ""

    if root.exists() and (Path(__file__).resolve().parents[2] / "config" / "settings.py").exists():
        from django.conf import settings

        if settings.MOCK_MODE:
            payload = _normalize_payload(_mock_payload(razdel_code, selected_doc_type_ids), razdel_code, selected_doc_type_ids)
            save_processing_json(project_id, "p4_doc_types_selection.json", {"selected_doc_type_ids": selected_doc_type_ids})
            save_processing_json(project_id, "p4_doc_plan.json", payload)
            save_processing_json(project_id, "p4b_doc_instances_v1.json", payload)
            if not (root / "02_processing" / "p4b_doc_instances_final.json").exists():
                save_processing_json(project_id, "p4b_doc_instances_final.json", payload)
            persist_run_json(project_id, "process_p4b", run_id, "uploaded_files.json", {"files": [], "input_texts": []})
            persist_run_artifact(project_id, "process_p4b", run_id, "raw_response.txt", json.dumps(payload, ensure_ascii=False))
            persist_run_json(project_id, "process_p4b", run_id, "p4_doc_plan.json", payload)
            persist_run_json(project_id, "process_p4b", run_id, "run_meta.json", {"process": "process_p4b", "model": model, "success": True, "mock_mode": True})
            return run_id, payload, []

    client = ResponsesClientV02(api_key=api_key or None)
    upload_paths, input_texts, skip_reasons = _build_request_inputs(project_id, razdel_code, input_files)
    upload_map: dict[str, str] = {}
    file_ids: list[str] = []
    excluded: list[dict] = []

    for p in upload_paths:
        try:
            rel = str(p.relative_to(root)).replace("\\", "/")
        except ValueError:
            rel = p.name
        try:
            fid = client.upload_file_bytes(p.name, p.read_bytes())
            upload_map[rel] = fid
            file_ids.append(fid)
        except Exception as e:
            err_str = str(e).lower()
            if "image_parse_error" in err_str or "unsupported image" in err_str or "invalid_request_error" in err_str:
                try:
                    if root in p.parents or p == root or str(p).startswith(str(root)):
                        _move_to_excluded(root, p, f"Ошибка загрузки: {e}")
                except (ValueError, TypeError):
                    pass
                excluded.append({"path": rel, "reason": str(e)[:300]})
            else:
                raise

    if not file_ids and not input_texts:
        excluded_msg = "; ".join(f"{x['path']}: {x['reason'][:80]}" for x in excluded) if excluded else ""
        raise ValueError(f"Не удалось собрать входы для Process 4. Исключённые файлы: {excluded_msg or 'нет'}.")

    system_prompt = load_prompt("01_system_v02")
    user_prompt = _load_p4_prompt() + f"\n\nproject_id={project_id}\nrazdel_code={razdel_code}\nrazdel_name={razdel.get('razdel_name', razdel_code)}\nselected_doc_type_ids={json.dumps(selected_doc_type_ids, ensure_ascii=False)}\nuser_comment={user_comment}"

    raw = ""
    last_error = ""
    payload: dict | None = None
    for attempt in range(1, 4):
        extra_texts = input_texts[:]
        if attempt > 1:
            extra_texts.append("Повтори ответ строго по JSON-контракту. Никакого текста вне JSON.")
        try:
            payload, raw = client.call_json_with_mixed_inputs(
                instructions=system_prompt,
                user_text=user_prompt,
                file_ids=file_ids,
                input_texts=extra_texts,
                model=model,
                timeout_s=1500,
            )
            is_valid, reason = validate_p4_plan_payload(payload)
            if is_valid:
                break
            last_error = reason
            persist_run_artifact(project_id, "process_p4b", run_id, f"raw_response_attempt_{attempt}.txt", raw)
            payload = None
        except Exception as exc:
            last_error = str(exc)
            persist_run_artifact(project_id, "process_p4b", run_id, f"raw_response_attempt_{attempt}.txt", raw or str(exc))
            payload = None

    if payload is None:
        persist_run_json(project_id, "process_p4b", run_id, "uploaded_files.json", {"uploaded": list(upload_map.keys()), "skip_reasons": skip_reasons, "input_texts_count": len(input_texts)})
        persist_run_artifact(project_id, "process_p4b", run_id, "raw_response.txt", raw)
        raise ValueError(f"Process 4: невалидный JSON после 3 попыток. Причина: {last_error}")

    payload = _normalize_payload(payload, razdel_code, selected_doc_type_ids)

    save_processing_json(project_id, "p4_doc_types_selection.json", {"selected_doc_type_ids": selected_doc_type_ids})
    save_processing_json(project_id, "p4_doc_plan.json", payload)
    save_processing_json(project_id, "p4b_doc_instances_v1.json", payload)
    if not (root / "02_processing" / "p4b_doc_instances_final.json").exists():
        save_processing_json(project_id, "p4b_doc_instances_final.json", payload)

    persist_run_json(project_id, "process_p4b", run_id, "uploaded_files.json", {"uploaded": upload_map, "skip_reasons": skip_reasons, "input_texts_count": len(input_texts)})
    persist_run_artifact(project_id, "process_p4b", run_id, "raw_response.txt", raw)
    persist_run_json(project_id, "process_p4b", run_id, "p4_doc_plan.json", payload)
    persist_run_json(project_id, "process_p4b", run_id, "run_meta.json", {"process": "process_p4b", "model": model, "success": True, "mock_mode": False, "excluded": excluded})
    if excluded:
        save_processing_json(project_id, "p4b_excluded_files.json", {"excluded": excluded, "run_id": run_id})
    return run_id, payload, excluded
