import re
import shutil
import zipfile
from pathlib import Path

from .project_storage import output_zip_path, project_root


def _load_openpyxl():
    try:
        from openpyxl import Workbook, load_workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "Для генерации XLSX требуется зависимость openpyxl. "
            "Установите: pip install -r requirements.txt"
        ) from exc
    return Workbook, load_workbook


CELL_RE = re.compile(r"^[A-Za-z]{1,3}[1-9]\d*$")


def _find_template(project_root_path: Path, template_ref: str) -> Path | None:
    name = (template_ref or "").strip()
    if not name:
        return None
    direct = (project_root_path / name).resolve()
    if direct.exists() and direct.is_file():
        return direct
    samples_root = project_root_path / "01_input" / "04_samples"
    if samples_root.exists():
        for p in samples_root.rglob("*"):
            if p.is_file() and p.name.lower() == Path(name).name.lower():
                return p
    return None


def _safe_sheet_name(name: str) -> str:
    cleaned = (name or "Лист1").strip()
    return cleaned[:31] or "Лист1"


def _write_xlsx_output(path: Path, payload: dict, project_root_path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    template_ref = (payload.get("template_ref") or "").strip()
    template_path = _find_template(project_root_path, template_ref)
    Workbook, load_workbook = _load_openpyxl()
    if template_path and template_path.suffix.lower() in {".xlsx", ".xlsm"}:
        shutil.copy2(template_path, path)
        wb = load_workbook(path)
    else:
        wb = Workbook()

    for fill in payload.get("fills") or []:
        if not isinstance(fill, dict):
            continue
        target = fill.get("target") if isinstance(fill.get("target"), dict) else {}
        sheet_name = _safe_sheet_name(str(target.get("sheet") or "Лист1"))
        cell = str(target.get("cell") or "").strip()
        if not CELL_RE.match(cell):
            continue
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
        ws[cell] = fill.get("value")
    wb.save(path)


def generate_from_fill_plan(project_id: str, fill_plan: dict) -> list[str]:
    root_path = project_root(project_id)
    root = root_path / "03_output"
    created: list[str] = []
    for output in fill_plan.get("outputs") or []:
        rel = (output.get("output_path") or "").strip().replace("\\", "/")
        if not rel:
            continue
        target = root / rel
        fmt = (output.get("format") or target.suffix.lstrip(".") or "xlsx").lower()
        if fmt in {"xlsx", "xlsm"}:
            try:
                _write_xlsx_output(target, output, root_path)
                created.append(rel)
            except RuntimeError as exc:
                if "openpyxl" not in str(exc).lower():
                    raise
                fallback = target.with_suffix(".txt")
                fallback.parent.mkdir(parents=True, exist_ok=True)
                fallback.write_text(str(output), encoding="utf-8")
                try:
                    rel_out = str(fallback.relative_to(root)).replace("\\", "/")
                except ValueError:
                    rel_out = fallback.name
                created.append(rel_out)
        else:
            # Minimal fallback: plain text when unknown format requested.
            # Kept explicit to avoid writing broken office files.
            target = target.with_suffix(".txt")
            target.parent.mkdir(parents=True, exist_ok=True)
            target.write_text(str(output), encoding="utf-8")
            try:
                rel_out = str(target.relative_to(root)).replace("\\", "/")
            except ValueError:
                rel_out = target.name
            created.append(rel_out)
    return created


def build_output_zip(project_id: str) -> Path | None:
    root = project_root(project_id) / "03_output"
    files = [p for p in root.rglob("*") if p.is_file() and p.name != "output.zip"]
    if not files:
        return None
    zip_path = output_zip_path(project_id)
    zip_path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for file_path in files:
            zf.write(file_path, file_path.relative_to(root))
    return zip_path
